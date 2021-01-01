#!/usr/bin/python

from __future__ import absolute_import, division, print_function

__metaclass__ = type

# import modules
import os
import openpyxl
from ansible.module_utils.basic import AnsibleModule
from openpyxl import Workbook


DOCUMENTATION = r"""
----
module: xls_write
Author: Sudarshan Vijaya Kumar
Version: 1.0
short_description: Write facts to excel spreadsheet
description:
- This module writes facts into a excel spreadsheet. The name of the spreadsheet 
    can be specified.
- The facts needs to be a list of dictonaries. It uses the key in the first item 
    in the list as headers to the excel spreadsheet.
- In the current state, it always returns as state changed because if it detects 
    that the workbook already has a spreadsheet with the same name, it will delete 
    the spreadsheet and create a new one with the new data.
- #TODO: Add support for checkmode and diff mode.
options:
    path:
        description: The path to the file that needs to be modified.
        type: path
        required: True
    workbook:
        description: The name of the excel spreadsheet that needs to be modified.
        type: str
        required: True
    worksheet: 
        description: The name of the work sheet that needs to be modified.
        type: str
        required: True
    data:
        description: The actual facts that need to be written in the file.
        type: list
        required: True
    create:
        description: If specified, the file will be created if it does not exist.
        type: bool
        required: False
        default: False
"""

EXAMPLES = r"""

    # Example data
    vars:
        data_list:
            - header1: value1
              header2: value2
              header3: value3
              
            - header1: another_value1
              header2: another_value2
              header3: another_value3
    
    #! Note: it is important and necessary to include the .xlsx extention in the
    #!       workbook name.
    - name: Write facts to spreadsheet
      xls_write:
        path: ./result
        workbook: workbook.xlsx
        worksheet: worksheet
        data: "{{ data_list }}"
        create: yes
"""


def write_xls_dict(module, dest, create, workbook, worksheet, data):
    """
    # This function does the following
    # - Check if destination path exists. Create the path if it does not exist
    #     provided the create parameter is set.
    # - Check if the destination file exists. Create the file it it does not exist
    #     provided the create parameter is set.
    # - If the destination file already exists, check to see there already exists
    #     a sheet with the same name as the one specified in the module. If yes,
    #     delete the existing worksheet.
    # - Create a new worksheet with the name as specified in the module.
    # - Take the first item from the list of data that is passed, extract the key,
    #     value pair and write the key as headers into the worksheet.
    # - Iterate through the list of dict that was passed as data and write to the
    #     spreadsheet.
    # - If a new workbook was created in the process, delete the default
    #     spreadseet that gets created and save the file before exiting the module.
    # TODO: Add support for dry run mode and diff computation.
    """
    # Variable to track if a change occured.
    changed = False
    # Variable to track if a workbook was created.
    work_book_created = False
    # Check if destination folder exists
    if not os.path.exists(dest):
        # Check if Create parameter is set to True
        if not create:
            # Module error message
            module.fail_json(
                msg=f"Destination folder '{dest}' does not exist"
                + "! Set create to 'True' to create the"
                + "destination folder."
            )
        # Try to create the destination folder.
        try:
            os.makedirs(dest)
            work_book = Workbook()
            work_book_created = True
            changed = True
        # raise exception and module error
        except Exception as err:
            module.fail_json(msg=f"Error creating {dest} ({err})")
    # If destination folder already exists, check to see if the destination file
    # exists
    else:
        if not os.path.isfile(f"{dest}/{workbook}"):
            # Check if Create parameter is set to True
            if not create:
                module.fail_json(
                    msg=f"Workbook '{workbook}' does not exist"
                    + "! Set Create to 'True' to create"
                    + "a new workbok."
                )
            # Try to open a workbook
            try:
                work_book = Workbook()
                work_book_created = True
                changed = True
            # raise exception and module error
            except Exception as err:
                module.fail_json(msg=f"Error creating {workbook} ({err})")
        # If the workbook already exists
        else:
            # try to open the workbook and read all data
            try:
                work_book = openpyxl.load_workbook(f"{dest}/{workbook}", data_only=True)
            # raise exception and module failure
            except Exception as err:
                module.fail_json(msg=f"Error creating {workbook} ({err})")
    # If no new workbook was created
    if not work_book_created:
        # loop through all sheet names and match with sheetname passed from
        # module. If matching sheet found, delete the sheet from workbook.
        # This provides a clean slate
        for sheet in work_book.sheetnames:
            if sheet == worksheet:
                work_book.remove(work_book[sheet])
                changed = True
    # create a new sheet with the name passed from module.
    work_sheet = work_book.create_sheet(f"{worksheet}")
    changed = True
    # Try to extract all key from the fist item in list. This will be used as
    # headers.
    try:
        headers = []
        for key, value in data[0].items():
            headers.append(str(key))
        work_sheet.append(headers)
    # if the data is not a list of directoried, the module will error out with
    # appropriate message
    except Exception as err:
        print(f"{err}: data must be a list of dictonaries.")
    # Loop through the list of data and write to spreadsheet.
    for entry in data:
        data_write = []
        for key, value in entry.items():
            data_write.append(str(value))
        work_sheet.append(data_write)
    # If a new workbook was created, delete the default sheet
    if work_book_created:
        work_book.remove(work_book["Sheet"])
    # save the workbook
    work_book.save(f"{dest}/{workbook}")
    # exit module with a status of changed and message of 'Done'
    module.exit_json(changed=changed, msg="Done!")


def main():
    # Define Module Parameters
    module = AnsibleModule(
        argument_spec=dict(
            path=dict(type="path", required=True),
            workbook=dict(type="str", required=True),
            worksheet=dict(type="str", required=True),
            data=dict(type="list", required=True),
            create=dict(type="bool", default=False),
        )
    )
    # Assign Module parameters
    params = module.params
    path = params["path"]
    workbook = params["workbook"]
    worksheet = params["worksheet"]
    data = params["data"]
    create = params["create"]
    # Ensure that the required parameters are passed and error out with an
    # appriopriate message if they are not.
    if path is None:
        module.fail_json(msg="path is required")
    if workbook is None:
        module.fail_json(msg="workbook is required")
    if worksheet is None:
        module.fail_json(msg="worksheet is required")
    if data is None:
        module.fail_json(msg="data is required")
    # Call function that handles the write operation
    write_xls_dict(module, path, create, workbook, worksheet, data)


# Invoke main function.
if __name__ == "__main__":
    main()
