#!/usr/bin/python

# import modules
import openpyxl
from ansible.module_utils.basic import *

"""
     Revision history:
     1 October 2020  |  1.0 - initial release
"""

DOCUMENTATION = r"""
---
module: xls_facts.py
author: Sudarshan Vijaya Kumar
version_added: "1.0"
short_description: Read an Excel .xlsx file and output Ansible facts
description:
    - Read the XLS file specified and output Ansible facts in the form of a list with each
      element in the list as a dictionary using the column header as the key and the contents
      of the cell as the value. A dictionary is created for each sheet,  in the format spreadsheet_SheetName.
 
requirements:
    - The openpyxl Python module must be installed on the Ansible host. This can be installed using pip:
      sudo pip install openpyxl  

options:
    src:
        description:
            - The name of the Excel spreadsheet
        required: true
   
"""

EXAMPLES = """

    Running the module from the command line:

      ansible localhost -m xls_facts -a src="example.xlsx" -M ~/ansible/library

   localhost | SUCCESS => {
    "ansible_facts": {
        "spreadsheet_Sheet1": [
            {
                "Hostname": "Switch-1",
                "Mgmt_ip": "10.0.0.1"
            },
            {
                "Hostname": "Switch-2",
                "Mgmt_ip": "10.0.0.2"
            },
            {
                "Hostname": "Switch-3",
                "Mgmt_ip": "10.0.0.3"
            }
        ],
        "spreadsheet_Sheet2": [
            {
                "Description": "To Spine-1",
                "Interface": "Ethernet1/1",
                "Interface_IP": "192.168.100.1/30"
            },
            {
                "Description": "To Spine-2",
                "Interface": "Ethernet1/2",
                "Interface_IP": "192.168.100.5/30"
            }
        ]
    },
    "changed": false

"""


# ---------------------------------------------------------------------------
# read_xls_dict
# ---------------------------------------------------------------------------


def read_xls_dict(input_file):
    "Read the XLS file and return as Ansible facts"
    spreadsheet = {}
    try:
        wb = openpyxl.load_workbook(input_file, data_only=True)
        for sheet in wb.sheetnames:
            ansible_sheet_name = f"spreadsheet_{sheet}"
            spreadsheet[ansible_sheet_name] = []
            current_sheet = wb[sheet]
            dict_keys = [
                current_sheet.cell(row=1, column=c).value
                for c in range(1, current_sheet.max_column + 1)
            ]
            for r in range(2, current_sheet.max_row + 1):
                temp_dict = {}
                for c in range(1, current_sheet.max_column + 1):
                    if value := current_sheet.cell(row=r, column=c).value:
                        temp_dict[dict_keys[c - 1]] = value
                    else:
                        value = ""
                spreadsheet[ansible_sheet_name].append(temp_dict)
    except IOError:
        return 1, f"IOError on input file:{input_file}"

    result = {"ansible_facts": spreadsheet}
    return (0, result)


def main():
    " "
    module = AnsibleModule(
        argument_spec=dict(src=dict(required=True)), add_file_common_args=True
    )

    code, response = read_xls_dict(module.params["src"])
    if code == 1:
        module.fail_json(msg=response)
    else:
        module.exit_json(**response)

    return code


if __name__ == "__main__":
    main()
#
